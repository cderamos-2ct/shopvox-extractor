#!/usr/bin/env python3
"""
build_workbook.py - Build a Google Sheets-ready workbook from ShopVox exports.

The workbook is designed for Drive import:
- one sheet per CSV export
- an index of source CSV/JSON files
- a PDF index with Drive search links generated from the transaction IDs

PDF hyperlinks intentionally use Drive search URLs instead of local file paths
so the workbook stays useful after the data has been copied into Google Drive.
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Dict, Iterable, List, Tuple
from urllib.parse import quote_plus

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


TRANSACTION_TYPES = {
    "quotes",
    "sales-orders",
    "invoices",
    "payments",
    "refunds",
}

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LINK_FONT = Font(color="0563C1", underline="single")


def drive_search_url(query: str) -> str:
    return f"https://drive.google.com/drive/search?q={quote_plus(query)}"


def safe_sheet_name(name: str, used: set[str]) -> str:
    raw = (
        name.replace("/", "__")
        .replace("\\", "__")
        .replace(":", "_")
        .replace("?", "_")
        .replace("*", "_")
        .replace("[", "_")
        .replace("]", "_")
    )
    raw = raw[:31].strip() or "Sheet"

    candidate = raw
    suffix = 2
    while candidate in used:
        # Keep the visible prefix stable and trim the tail when de-duplicating.
        tail = f"_{suffix}"
        candidate = f"{raw[:31 - len(tail)]}{tail}"
        suffix += 1

    used.add(candidate)
    return candidate


def iter_files(root: Path, suffix: str) -> Iterable[Path]:
    for path in sorted(root.rglob(f"*{suffix}")):
        if path.is_file():
            yield path


def read_csv_rows(csv_path: Path) -> Tuple[List[str], List[List[str]]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            header = next(reader)
        except StopIteration:
            return [], []
        rows = [row for row in reader]
    return header, rows


def load_json(json_path: Path) -> dict:
    with json_path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def display_name_for_record(record: dict, type_name: str) -> str:
    data = record.get("data") or {}
    candidates = {
        "quotes": ["quoteNumber", "number", "quote_no", "name"],
        "sales-orders": ["workOrderNumber", "workOrderNo", "number", "name"],
        "invoices": ["invoiceNumber", "number", "name"],
        "payments": ["paymentNumber", "number", "name"],
        "refunds": ["refundNumber", "number", "name"],
    }.get(type_name, ["number", "name", "id"])

    for key in candidates:
        value = data.get(key)
        if value not in (None, ""):
            return str(value)
    return str(record.get("id") or data.get("id") or "")


def write_header_row(ws, headers: List[str]) -> None:
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top")
    ws.freeze_panes = "A2"


def autosize_columns(ws, max_width: int = 60) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            value = str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), min(len(value), max_width))

    for col_idx, width in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(width + 2, max_width + 2)


def add_index_sheet(wb: Workbook, data_dir: Path, csv_files: List[Path], json_files: List[Path]) -> None:
    ws = wb.create_sheet("Source_Index")
    headers = ["kind", "relative_path", "file_name", "drive_search_link", "sheet_name"]
    write_header_row(ws, headers)

    row = 2
    for path in csv_files + json_files:
        rel = path.relative_to(data_dir).as_posix()
        kind = "csv" if path.suffix.lower() == ".csv" else "json"
        sheet_name = path.stem
        ws.cell(row=row, column=1, value=kind)
        ws.cell(row=row, column=2, value=rel)
        ws.cell(row=row, column=3, value=path.name)
        link_cell = ws.cell(row=row, column=4, value=drive_search_url(path.name))
        link_cell.hyperlink = drive_search_url(path.name)
        link_cell.font = LINK_FONT
        ws.cell(row=row, column=5, value=sheet_name if kind == "csv" else "")
        row += 1

    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)


def add_csv_sheet(wb: Workbook, csv_path: Path, data_dir: Path, used_names: set[str]) -> str:
    rel = csv_path.relative_to(data_dir).with_suffix("")
    sheet_name = safe_sheet_name("__".join(rel.parts), used_names)
    ws = wb.create_sheet(sheet_name)
    type_name = csv_path.stem

    header, rows = read_csv_rows(csv_path)
    if not header:
        ws["A1"] = "Empty CSV file"
        return sheet_name

    headers = list(header)
    is_transaction_type = type_name in TRANSACTION_TYPES
    if is_transaction_type:
        headers.append("pdf_link")

    write_header_row(ws, headers)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        if is_transaction_type:
            link_cell = ws.cell(
                row=row_idx,
                column=len(headers),
                value=f'=HYPERLINK("https://drive.google.com/drive/search?q="&ENCODEURL($A{row_idx}&".pdf"),"PDF")',
            )
            link_cell.font = LINK_FONT

    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)
    return sheet_name


def add_pdf_index(wb: Workbook, data_dir: Path, json_files: List[Path]) -> None:
    ws = wb.create_sheet("PDF_Index")
    headers = [
        "type",
        "record_id",
        "display_name",
        "pdf_file_name",
        "pdf_drive_search_link",
        "source_json",
    ]
    write_header_row(ws, headers)

    row = 2
    for json_path in json_files:
        type_name = json_path.stem
        if type_name not in TRANSACTION_TYPES:
            continue

        payload = load_json(json_path)
        records = payload.get("records") or []
        source_link = drive_search_url(json_path.name)
        for record in records:
            record_id = str(record.get("id") or "")
            if not record_id:
                continue
            pdf_name = f"{record_id}.pdf"
            pdf_link = drive_search_url(pdf_name)
            ws.cell(row=row, column=1, value=type_name)
            ws.cell(row=row, column=2, value=record_id)
            ws.cell(row=row, column=3, value=display_name_for_record(record, type_name))
            ws.cell(row=row, column=4, value=pdf_name)
            link_cell = ws.cell(row=row, column=5, value=pdf_link)
            link_cell.hyperlink = pdf_link
            link_cell.font = LINK_FONT
            source_cell = ws.cell(row=row, column=6, value=source_link)
            source_cell.hyperlink = source_link
            source_cell.font = LINK_FONT
            row += 1

    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)


def add_overview_sheet(wb: Workbook, data_dir: Path, csv_files: List[Path], json_files: List[Path]) -> None:
    ws = wb.create_sheet("Overview", 0)
    ws["A1"] = "ShopVox Extraction Index"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = "Generated from the extracted CSV and JSON files, with Drive-search links for PDFs."
    ws["A4"] = "Data directory"
    ws["B4"] = str(data_dir)
    ws["A5"] = "CSV files"
    ws["B5"] = len(csv_files)
    ws["A6"] = "JSON files"
    ws["B6"] = len(json_files)
    ws["A8"] = "Open Source_Index for file links"
    ws["B8"] = "Source_Index"
    ws["B8"].hyperlink = "#Source_Index!A1"
    ws["B8"].font = LINK_FONT
    ws["A9"] = "Open PDF_Index for transaction PDFs"
    ws["B9"] = "PDF_Index"
    ws["B9"].hyperlink = "#PDF_Index!A1"
    ws["B9"].font = LINK_FONT
    ws["A11"] = "Notes"
    ws["A12"] = "PDF links use Drive search by record ID (filename = <record_id>.pdf)."
    ws["A13"] = "Import this workbook into Google Sheets to keep the hyperlinks clickable in Drive."

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90


def build_workbook(data_dir: Path, output_path: Path) -> Path:
    if not data_dir.exists():
        raise FileNotFoundError(f"Data directory not found: {data_dir}")

    csv_files = list(iter_files(data_dir, ".csv"))
    json_files = list(iter_files(data_dir, ".json"))

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    add_overview_sheet(wb, data_dir, csv_files, json_files)
    add_index_sheet(wb, data_dir, csv_files, json_files)

    used_names = {"Overview", "Source_Index", "PDF_Index"}
    for csv_path in csv_files:
        add_csv_sheet(wb, csv_path, data_dir, used_names)

    add_pdf_index(wb, data_dir, json_files)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a ShopVox extraction workbook.")
    parser.add_argument(
        "--data-dir",
        default="./shopvox-data",
        help="Root directory containing the extracted CSV and JSON files.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Workbook output path. Defaults to <data-dir>/ShopVox-Extraction-Index.xlsx",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    data_dir = Path(args.data_dir).expanduser().resolve()
    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else data_dir / "ShopVox-Extraction-Index.xlsx"
    )

    written = build_workbook(data_dir, output_path)
    print(f"Wrote {written}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
